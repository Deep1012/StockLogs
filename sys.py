import streamlit as st
import pandas as pd
import requests
from datetime import datetime
import openpyxl
import os

API_KEY = "6QZ3IT21J6UL0M24"

# Load the list of Indian stocks from the provided Excel file
@st.cache_data
def load_indian_stocks():
    file_path = "C:\\Users\\DELL\\Desktop\\StockLog\\data.xlsx" # Path to the uploaded Excel file
    df = pd.read_excel(file_path)
    stocks_dict = df.set_index("Symbol")["Company Name"].to_dict()
    return stocks_dict

indian_stocks = load_indian_stocks()

# Function to get real-time stock price
def get_stock_price(symbol):
    url = f"https://www.alphavantage.co/query?function=GLOBAL_QUOTE&symbol={symbol}.BSE&apikey={API_KEY}"
    try:
        response = requests.get(url)
        data = response.json()
        
        if "Global Quote" in data and "05. price" in data["Global Quote"]:
            return float(data["Global Quote"]["05. price"])
        else:
            return None
    except Exception as e:
        st.error(f"Error fetching stock price: {str(e)}")
        return None

# Function to validate and get exact stock symbol
def validate_and_get_stock_symbol(keyword):
    keyword = keyword.upper()
    if keyword in indian_stocks:
        return keyword, indian_stocks[keyword]
    
    # If not found, try partial matching
    matches = [symbol for symbol in indian_stocks if keyword in symbol or keyword in indian_stocks[symbol].upper()]
    if matches:
        return matches[0], indian_stocks[matches[0]]
    
    return None, None

# Function to log data to a new date-wise Excel file
def log_to_excel(name, price, quantity, order_type, total_price):
    # Ensure the Logs folder exists
    logs_folder = "C:\\Users\\DELL\\Desktop\\ERP\\Logs"
    if not os.path.exists(logs_folder):
        os.makedirs(logs_folder)

    # Create a filename based on the current date in the Logs folder
    today = datetime.now().strftime("%Y-%m-%d")
    filename = os.path.join(logs_folder, f"stock_log_{today}.xlsx")
    
    if not os.path.exists(filename):
        df = pd.DataFrame(columns=["Name", "Date and Time", "Price", "Quantity", "Order Type", "Total Price"])
        df.to_excel(filename, index=False)
    
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([name, now, price, quantity, order_type, total_price])
    
    # Calculate and update total buy and sell for the day
    total_buy = sum(sheet[f"F{i}"].value for i in range(2, sheet.max_row + 1) if sheet[f"E{i}"].value == "Buy")
    total_sell = sum(sheet[f"F{i}"].value for i in range(2, sheet.max_row + 1) if sheet[f"E{i}"].value == "Sell")
    
    sheet["H1"] = "Total Buy"
    sheet["H2"] = total_buy
    sheet["I1"] = "Total Sell"
    sheet["I2"] = total_sell
    
    wb.save(filename)

# Streamlit app
st.title("Stock Management System")

# Sidebar for stock price search
st.sidebar.header("Stock Price Search")
search_stock = st.sidebar.text_input("Enter stock name or symbol to search:")
if st.sidebar.button("Search Stock Price"):
    if search_stock:
        symbol, name = validate_and_get_stock_symbol(search_stock)
        if symbol:
            price = get_stock_price(symbol)
            if price:
                st.sidebar.success(f"Real-time price for {name} ({symbol}): ₹{price:.2f}")
            else:
                st.sidebar.error("Unable to fetch real-time price. Please try again.")
        else:
            st.sidebar.error("Stock not found. Please check the name or symbol and try again.")

# Main content
st.header("Log Stock Transaction")

# Input fields
stock_name = st.text_input("Enter stock name or symbol:")
order_type = st.selectbox("Order type:", ["Buy", "Sell"])
quantity = st.number_input("Quantity of shares:", min_value=1, step=1)
price = st.number_input("Price per share:", min_value=0.01, step=0.01)

# Calculate total price
total_price = price * quantity
st.info(f"Total transaction value: ₹{total_price:.2f}")

# Submit order
if st.button("Submit Order"):
    if stock_name and quantity and price > 0:
        symbol, name = validate_and_get_stock_symbol(stock_name)
        if symbol:
            log_to_excel(f"{name} ({symbol})", price, quantity, order_type, total_price)
            st.success(f"Order logged: {quantity} shares of {name} ({symbol}) {order_type.lower()}n at ₹{price:.2f} per share. Total: ₹{total_price:.2f}")
        else:
            st.error("Invalid stock name or symbol. Please enter a correct stock name or symbol from the Indian stock market.")
    else:
        st.warning("Please enter valid stock name, quantity, and price.")

# Display logged data
if st.checkbox("Show Logged Data"):
    today = datetime.now().strftime("%Y-%m-%d")
    filename = os.path.join("C:\\Users\\DELL\\Desktop\\ERP\\Logs", f"stock_log_{today}.xlsx")
    if os.path.exists(filename):
        df = pd.read_excel(filename)
        
        # Remove unnamed columns and drop rows that are entirely empty
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        df.dropna(how='all', inplace=True)
        
        # Replace None/NaN values with empty strings
        df.fillna('', inplace=True)
        
        st.dataframe(df)
    else:
        st.info("No data logged yet.")
